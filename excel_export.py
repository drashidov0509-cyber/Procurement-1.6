"""
Экспорт результатов поиска в Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def export_to_excel(result: dict, country_name: str, save_path: str):
    """Создаёт Excel-файл с результатами"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # Стили
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    bold = Font(name="Calibri", size=11, bold=True)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Заголовок документа
    ws["A1"] = f"ЗАКУПОЧНАЯ ВЕДОМОСТЬ — {country_name}, {result.get('region', '—')}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells("A1:K1")

    ws["A2"] = (
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   "
        f"Валюта: {result.get('currency', 'UZS')}   |   "
        f"НДС: {result.get('vat_rate', 12)}%"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells("A2:K2")

    # Шапка таблицы
    headers = [
        "№", "Наименование", "Параметры", "Ед.", "Кол-во",
        "Категория", "Поставщик", "Адрес/контакты", "Источник",
        f"Цена/ед ({result.get('currency', '')})",
        f"Итого с НДС ({result.get('currency', '')})",
    ]
    row_idx = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Данные
    tier_names = {"cheap": "Дешёвый", "mid": "Средний", "exp": "Дорогой"}
    tier_colors = {"cheap": "D5F5D5", "mid": "FFF4CC", "exp": "FFD9D9"}

    row_idx = 5
    for it in result.get("rows", []):
        for s in it.get("suppliers", []):
            row = [
                it["item_num"],
                it["item_name"],
                it.get("item_param", ""),
                it.get("item_unit", ""),
                it["item_qty"],
                tier_names.get(s["tier"], s["tier"]),
                s["supplier_name"],
                s.get("supplier_address", ""),
                s.get("source", ""),
                s["price_per_unit"],
                s["total"],
            ]
            for col_idx, v in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col_idx == 6:  # категория
                    cell.fill = PatternFill("solid", fgColor=tier_colors.get(s["tier"], "FFFFFF"))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx in (10, 11):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            row_idx += 1
        # Пустая строка между позициями
        row_idx += 1

    # Итоги
    row_idx += 1
    ws.cell(row=row_idx, column=10, value="ИТОГО мин (дешёвый):").font = bold
    ws.cell(row=row_idx, column=11, value=result.get("total_min", 0)).font = bold
    ws.cell(row=row_idx, column=11).number_format = '#,##0.00'

    ws.cell(row=row_idx + 1, column=10, value="ИТОГО макс (дорогой):").font = bold
    ws.cell(row=row_idx + 1, column=11, value=result.get("total_max", 0)).font = bold
    ws.cell(row=row_idx + 1, column=11).number_format = '#,##0.00'

    # Ширина колонок
    widths = [4, 28, 22, 7, 8, 11, 28, 32, 18, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Высота шапки
    ws.row_dimensions[4].height = 32

    # Заморозить верхушку
    ws.freeze_panes = "A5"

    # ───── Лист 2: Шаблон импорта ─────
    ws2 = wb.create_sheet("Шаблон импорта")
    ws2.append(["Наименование ТМЦ", "Параметры", "Ед.изм.", "Объём"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="E0E8F5")
    ws2.append(["Цемент М400", "мешок 50 кг, ГОСТ 31108", "шт", 100])
    ws2.append(["Труба ПВХ", "d32мм, L=3м", "шт", 50])
    ws2.append(["Кабель ВВГ", "3×2.5мм²", "м", 500])
    for letter, w in zip("ABCD", [32, 28, 10, 10]):
        ws2.column_dimensions[letter].width = w

    wb.save(save_path)
