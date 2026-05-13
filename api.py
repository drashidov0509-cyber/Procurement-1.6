"""
Procurement Global v1.6 — API мост Python ↔ JavaScript
"""
import os, sys, json, asyncio, subprocess, re, time, webview
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from scrapers import search_all_sources
from classifier import classify_suppliers
from excel_export import export_to_excel

VAT_RATES = {
    "UZ":12,"AZ":18,"KZ":12,"KG":12,"TJ":18,"TM":15,
    "RU":20,"TR":20,"CN":13,"DE":19,"US":0,"AE":5,
    "GB":20,"PL":23,"GE":18,"AM":20,
}

def _pip(pkg):
    subprocess.check_call(
        [sys.executable,"-m","pip","install","-q",pkg],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )


class API:
    def __init__(self):
        self._window = None

    def _set_window(self, w):
        self._window = w

    def ping(self):
        return {"status":"ok","version":"1.6.0"}

    # ── ПОИСК ПОСТАВЩИКОВ ──────────────────────────────────────────
    def search_suppliers(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        try:
            country  = payload.get("country","UZ")
            region   = payload.get("region","")
            currency = payload.get("currency","UZS")
            items    = payload.get("items",[])
            if not items:
                return {"error":"Список позиций пуст"}

            vat_rate = VAT_RATES.get(country, 12)
            rows = []
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                for idx, item in enumerate(items, 1):
                    self._notify(idx, len(items), item.get("name",""))
                    query = item.get("name","")
                    if item.get("param"):
                        query += f" {item['param']}"
                    listings = loop.run_until_complete(
                        search_all_sources(query, country=country, region=region)
                    )
                    suppliers = classify_suppliers(
                        listings, item.get("qty",1), query=query
                    )
                    rows.append({
                        "item_num":   idx,
                        "item_name":  item.get("name",""),
                        "item_param": item.get("param",""),
                        "item_unit":  item.get("unit","шт"),
                        "item_qty":   item.get("qty",1),
                        "suppliers":  [s.dict() for s in suppliers],
                    })
            finally:
                loop.close()

            total_min = sum(min((s["total"] for s in r["suppliers"]),default=0) for r in rows)
            total_max = sum(max((s["total"] for s in r["suppliers"]),default=0) for r in rows)
            return {
                "country":country,"region":region,"currency":currency,
                "vat_rate":vat_rate,"rows":rows,
                "total_min":round(total_min,2),"total_max":round(total_max,2),
                "created_at":datetime.now().isoformat(),
            }
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error":f"Ошибка поиска: {e}"}

    def _notify(self, cur, total, name):
        if self._window:
            try:
                self._window.evaluate_js(
                    f"window.onSearchProgress&&window.onSearchProgress("
                    f"{cur},{total},{json.dumps(name)})"
                )
            except Exception:
                pass

    # ── ИМПОРТ EXCEL ───────────────────────────────────────────────
    def import_excel(self) -> Dict[str, Any]:
        try:
            if not self._window:
                return {"error":"Окно недоступно"}
            files = self._window.create_file_dialog(
                dialog_type=webview.OPEN_DIALOG, allow_multiple=False,
                file_types=("Excel (*.xlsx;*.xls)",),
            )
            if not files:
                return {"cancelled":True}
            path = files[0] if isinstance(files,(list,tuple)) else files

            try:
                from openpyxl import load_workbook
            except ImportError:
                _pip("openpyxl"); from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            items = []
            for row in list(ws.iter_rows(values_only=True))[1:]:
                if not row or not row[0]: continue
                name  = str(row[0]).strip()
                param = str(row[1]).strip() if len(row)>1 and row[1] else ""
                unit  = str(row[2]).strip() if len(row)>2 and row[2] else "шт"
                try:   qty = float(row[3]) if len(row)>3 and row[3] else 1.0
                except: qty = 1.0
                if name and qty>0:
                    items.append({"name":name,"param":param,"unit":unit,"qty":qty})
            wb.close()
            return {"ok":True,"items":items,"count":len(items)}
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error":str(e)}

    # ── ИМПОРТ PDF ─────────────────────────────────────────────────
    def import_pdf(self) -> Dict[str, Any]:
        try:
            if not self._window:
                return {"error": "Окно недоступно"}
            files = self._window.create_file_dialog(
                dialog_type=webview.OPEN_DIALOG, allow_multiple=False,
                file_types=("PDF (*.pdf)",),
            )
            if not files:
                return {"cancelled": True}
            path = files[0] if isinstance(files, (list, tuple)) else files

            from pdf_parser import parse_pdf
            items = parse_pdf(path)

            if not items:
                return {"error":
                    "Не удалось распознать позиции ТМЦ в PDF.\n"
                    "Убедитесь что PDF содержит таблицу:\n"
                    "Наименование | Параметры | Ед.изм. | Кол-во"}
            return {"ok": True, "items": items, "count": len(items)}
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error": f"Ошибка чтения PDF: {e}"}

    # ── ЭКСПОРТ EXCEL ──────────────────────────────────────────────
    def export_excel(self, result: Dict[str,Any], country_name: str,
                     save_path: str = None) -> Dict[str,Any]:
        try:
            if not save_path and self._window:
                files = self._window.create_file_dialog(
                    dialog_type=webview.SAVE_DIALOG,
                    save_filename=f"procurement_{country_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    file_types=("Excel (*.xlsx)",),
                )
                if not files: return {"cancelled":True}
                save_path = files[0] if isinstance(files,(list,tuple)) else files
            export_to_excel(result, country_name, save_path)
            return {"ok":True,"path":save_path}
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error":str(e)}

    # ── ЭКСПОРТ PDF ────────────────────────────────────────────────
    def export_pdf(self, result: Dict[str,Any], country_name: str) -> Dict[str,Any]:
        try:
            if not self._window:
                return {"error":"Окно недоступно"}
            files = self._window.create_file_dialog(
                dialog_type=webview.SAVE_DIALOG,
                save_filename=f"procurement_{country_name}_{datetime.now().strftime('%Y%m%d')}.pdf",
                file_types=("PDF (*.pdf)",),
            )
            if not files: return {"cancelled":True}
            save_path = files[0] if isinstance(files,(list,tuple)) else files

            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import (SimpleDocTemplate, Table,
                    TableStyle, Paragraph, Spacer)
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import mm
            except ImportError:
                _pip("reportlab")
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import (SimpleDocTemplate, Table,
                    TableStyle, Paragraph, Spacer)
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import mm

            doc = SimpleDocTemplate(save_path, pagesize=A4,
                leftMargin=15*mm, rightMargin=15*mm,
                topMargin=15*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            story  = []

            # Заголовок
            story.append(Paragraph(
                f"<b>Закупочная ведомость — {country_name}, {result.get('region','')}</b>",
                ParagraphStyle("t",parent=styles["Title"],fontSize=13,spaceAfter=4)
            ))
            story.append(Paragraph(
                f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')} &nbsp;|&nbsp; "
                f"Валюта: {result.get('currency','')} &nbsp;|&nbsp; "
                f"НДС: {result.get('vat_rate','')}%",
                ParagraphStyle("s",parent=styles["Normal"],fontSize=9,
                               textColor=colors.grey, spaceAfter=12)
            ))

            tn = {"cheap":"Дешёвый","mid":"Средний","exp":"Дорогой"}
            tc = {"cheap":colors.HexColor("#d5f5d5"),
                  "mid":  colors.HexColor("#fff4cc"),
                  "exp":  colors.HexColor("#ffd9d9")}
            cur = result.get("currency","")

            for it in result.get("rows",[]):
                label = f"{it['item_num']}. {it['item_name']}"
                if it.get("item_param"): label += f" ({it['item_param']})"
                label += f" — {it['item_qty']} {it['item_unit']}"
                story.append(Paragraph(f"<b>{label}</b>", styles["Normal"]))
                story.append(Spacer(1,3*mm))

                tdata = [["Категория","Поставщик","Адрес",
                          f"Цена/ед ({cur})",f"Итого ({cur})"]]
                for s in it.get("suppliers",[]):
                    tdata.append([
                        tn.get(s["tier"],s["tier"]),
                        (s["supplier_name"]  or "")[:40],
                        (s["supplier_address"] or "")[:40],
                        f"{s['price_per_unit']:,.2f}",
                        f"{s['total']:,.2f}",
                    ])

                tbl = Table(tdata, colWidths=[25*mm,50*mm,50*mm,28*mm,28*mm])
                ts  = TableStyle([
                    ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1f2d45")),
                    ("TEXTCOLOR", (0,0),(-1,0),colors.white),
                    ("FONTSIZE",  (0,0),(-1,-1),8),
                    ("ALIGN",     (3,0),(4,-1),"RIGHT"),
                    ("GRID",      (0,0),(-1,-1),0.5,colors.HexColor("#cccccc")),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),
                     [colors.white, colors.HexColor("#f8f9fa")]),
                ])
                for ri, s in enumerate(it.get("suppliers",[]),1):
                    ts.add("BACKGROUND",(0,ri),(0,ri),tc.get(s["tier"],colors.white))
                tbl.setStyle(ts)
                story.append(tbl)
                story.append(Spacer(1,6*mm))

            # Итоговая строка
            story.append(Paragraph(
                f"<b>Итого мин: {result.get('total_min',0):,.2f} {cur} &nbsp;&nbsp; "
                f"Итого макс: {result.get('total_max',0):,.2f} {cur}</b>",
                styles["Normal"]
            ))

            doc.build(story)
            return {"ok":True,"path":save_path}
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error":str(e)}

    # ── СКРИНШОТ ───────────────────────────────────────────────────
    def take_screenshot(self) -> Dict[str,Any]:
        try:
            try:
                from PIL import ImageGrab
            except ImportError:
                _pip("pillow"); from PIL import ImageGrab

            if self._window:
                files = self._window.create_file_dialog(
                    dialog_type=webview.SAVE_DIALOG,
                    save_filename=f"procurement_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                    file_types=("PNG (*.png)",),
                )
                if not files: return {"cancelled":True}
                save_path = files[0] if isinstance(files,(list,tuple)) else files
            else:
                save_path = str(
                    Path.home()/f"procurement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                )

            time.sleep(0.4)          # дать диалогу закрыться
            img = ImageGrab.grab()
            img.save(save_path,"PNG")
            return {"ok":True,"path":save_path}
        except Exception as e:
            import traceback; traceback.print_exc()
            return {"error":f"Ошибка скриншота: {e}"}
